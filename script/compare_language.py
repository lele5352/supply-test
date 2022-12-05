import json
import re
from openpyxl import load_workbook
from utils.log_handler import logger as log

ZH_CN_PATH = '/Users/linzhongjie/Downloads/messages_zh_CN.properties'
FR_FR_PATH = '/Users/linzhongjie/Downloads/messages_fr_FR.properties'
EN_US_PATH = '/Users/linzhongjie/Downloads/messages_en_US.properties'
# TRANS_FILE_PATH = '/Users/linzhongjie/Downloads/国际化文档.xlsx'
TRANS_FILE_PATH = '/Users/linzhongjie/Downloads/国际化语言包(内部).xlsx'

reg = "[^0-9A-Za-z\u4e00-\u9fa5]"


def handle_str(raw_str: str):
    """
    处理字符
    """
    return raw_str.replace('\u00A0', '').replace(',', '，')\
        .replace("!", "！").replace(":", "：").replace(" ", "").replace('\\', '')


def get_properties_data(file_path: str) -> dict:
    """
    解析配置文件数据，返回 dict
    :param file_path: 文件路径
    """
    properties = {}

    with open(file_path, 'r', encoding='utf-8') as config_file:

        for line in config_file:
            if line.find('=') > 0 and line.startswith('rest.i18n'):
                # 过滤掉导出的配置
                strs = line.replace('\n', '').split('=', 1)  # 以 = 分割
                properties[strs[0]] = strs[1]

    return properties


def get_unduplicated_data(file_path: str) -> dict:
    """
    解析国际化翻译文档（已合并前后端文案到同一个sheet,不需要程序再处理）
    """
    work_book = load_workbook(file_path)
    trans_data = {}
    target_sheet = work_book['中文去重翻译(客服)']

    for row in target_sheet.rows:
        # 过滤可能存在的空行
        if row[0].value is None:
            continue

        try:
            trans_data[handle_str(row[0].value)] = {
                "us": row[1].value,
                "fr": row[2].value
            }
        except TypeError:
            log.error("翻译文案可能缺失，检查: %s" % row[0].value)
            continue

    return trans_data


def get_trans_data(file_path: str) -> dict:
    """
    解析国际化翻译文档（未将前后端文案合并到同一个sheet,需要做合并去重）
    """
    work_book = load_workbook(file_path)
    front_map, backend_map = {}, {}

    front_sheet = work_book["前端中文去重(客服)"]
    backend_sheet = work_book['后端中文去重(客服)']

    for front in front_sheet.rows:

        if front[0].value is None:
            continue

        front_map[handle_str(front[0].value)] = {
            "us": front[3].value,
            "fr": front[4].value
        }

    for backend in backend_sheet.rows:

        if backend[0].value is None:
            continue

        try:
            backend_map[handle_str(backend[0].value)] = {
                "us": backend[2].value,
                "fr": backend[3].value
            }
        except TypeError:
            log.error('翻译文案可能缺失，检查: %s' % backend[0].value)
            continue

    backend_map.update(front_map)

    return backend_map


def bulid_lang_dict() -> dict:
    """
    以中文为索引，将3个语言配置合成一份，返回 dict
    """
    result_map = {}

    ch_data = get_properties_data(ZH_CN_PATH)
    us_data = get_properties_data(EN_US_PATH)
    fr_data = get_properties_data(FR_FR_PATH)

    for k, v in ch_data.items():
        v_k = handle_str(v)

        if v_k not in result_map:
            # 相同中文文案的，只保留一个

            result_map[v_k] = {
                "code": k
            }

            if k in us_data:
                result_map[v_k]['us'] = us_data[k]
            if k in fr_data:
                result_map[v_k]['fr'] = fr_data[k]

    return result_map


def do_compare() -> list:
    """
    对比翻译文档结果和配置，
    list 返回对比结果不一致的数据
    """
    wrong_list = []

    # backend_rs = get_trans_data(TRANS_FILE_PATH)
    backend_rs = get_unduplicated_data(TRANS_FILE_PATH)

    lang_rs = bulid_lang_dict()

    for k, v in lang_rs.items():

        if k not in backend_rs:
            wrong_list.append(k)
            log.error("code：%s ，文案: %s ，在线文档中不存在" % (v['code'], k))
            continue

        if re.sub(reg, '', v['us']) != re.sub(reg, '', backend_rs[k]['us']):
            wrong_list.append(k)
            log.error("code：%s ，文案：%s ，读取配置英语文案为 %s \n "
                      "在线文档结果为：%s" % (v['code'], k, v['us'], backend_rs[k]['us']))
            continue
        else:
            if re.sub(reg, '', v['fr']) != re.sub(reg, '', backend_rs[k]['fr']):
                wrong_list.append(k)
                log.error("code：%s ，文案：%s ，读取配置法语文案为 %s \n "
                          "在线文档结果为：%s" % (v['code'], k, v['fr'], backend_rs[k]['fr']))
                continue

    return wrong_list


if __name__ == '__main__':

    rs = do_compare()

    if rs:
        print('输出校验不通过列表：%s' % json.dumps(rs, ensure_ascii=False))
    else:
        print("校验完成，未检查到差异项")


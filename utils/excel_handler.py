import concurrent.futures
import json
import time

import openpyxl


class ExcelTool:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = openpyxl.load_workbook(self.filename)

    def multi_read2dict(self, key_map=None, sheet_name=None):
        """
        多线程读取excel,根据配置key,将每行数据映射成字典,最终返回list
        """
        if not sheet_name:
            sheet_name = self.workbook.sheetnames[0]

        if not key_map:
            key_map = {}

        sheet = self.workbook[sheet_name]

        # 获取列标题
        headers = [cell.value for cell in sheet[1]]

        # 初始化一个列表用于存储映射后的字典
        data = []

        def process_row(row):
            row_dict = {}
            # 将每个单元格的数据映射到对应的键
            for header, cell_value in zip(headers, row):
                key = key_map.get(header, header)  # 使用配置中的映射或默认列标题
                row_dict[key] = cell_value
            data.append(row_dict)

        # 使用线程池并行处理数据
        with concurrent.futures.ThreadPoolExecutor() as executor:
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            executor.map(process_row, rows)

        return data

    def read_data(self, sheet_name=None, start_row=1, jsonfy=False):
        if sheet_name is None:
            sheet_name = self.workbook.sheetnames[0]
        sheet = self.workbook[sheet_name]

        data = []
        for row in sheet.iter_rows(min_row=start_row, values_only=True):
            data.append([json.loads(cell) if cell else {} for cell in row] if jsonfy else row)
        return data

    def write_data(self, data, sheet_name=None, start_row=None):
        if sheet_name is None:
            sheet_name = 'Sheet1'

        if sheet_name not in self.workbook.sheetnames:
            self.workbook.create_sheet(sheet_name)
        sheet = self.workbook[sheet_name]

        if start_row is None:
            start_row = sheet.max_row + 1

        for row_data in data:
            sheet.append(row_data)

        self.workbook.save(self.filename)


if __name__ == '__main__':
    # data = ExcelTool("../test_data/transfer_test_data.xlsx")
    # fhc_cp_other_in_data = ExcelTool("../test_data/ims_test_data.xlsx").read_data("fhc_cp_other_in",2)
    start_time = time.time()
    print(f"开始时间{start_time}")
    fhc_cp_other_in_data = ExcelTool("../test_data/Braintree_Popi.xlsx").multi_read2dict()
    # print(fhc_cp_other_in_data)
    exec_time = time.time() - start_time
    print(f"执行耗时{exec_time}秒")

